#======USER TO FILL===========
url = "https://fsm.sg.formulasquare.com/fsm_api/wawaji_cms/"
#=============================

import time
import openpyxl
import streamlit as st
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options

options = Options()
options.headless = True

st.title("5G Claw Machine CMS - input coupon code from excel file")
st.write("Please use the form below to enter the codes into the CMS system. Note: the form below accepts coupon codes that are whole numbers only.")
st.write("If at any point you encounter an error message, this could be due to the following: 1) username and/or password is incorrrect; or 2) coupon code already exists in the CMS system.")

uploaded_file = st.file_uploader("Upload file", type=['xlsx'])

form = st.form("coupon_codes_excel_form")

username = form.text_input("Please enter the CMS username")
password = form.text_input("Please enter the CMS password")
row_start = form.text_input("Row number that corresponds to where the codes begin (e.g. if a header row exists, please key in 2)")
col_num = form.text_input("Column number containing codes (e.g. key in 1 if coupon codes are in column A, 2 if in column B, etc.)")
submit = form.form_submit_button("Add coupon codes to CMS")

if uploaded_file is not None:
    if submit:
        row_start = int(row_start)
        col_num = int(col_num)
        st.info("Running. Please do NOT click on the button again.")

        #open workbook
        wrkbk = openpyxl.load_workbook(uploaded_file)
        sh = wrkbk.active

        web = webdriver.Chrome(ChromeDriverManager(options=options)
        web.get(url)
        time.sleep(3)

        #Login
        username_field = web.find_element("xpath", '/html/body/div/div/main/div/div/div/div/form/div[1]/input')
        username_field.send_keys(username)
        password_field = web.find_element("xpath", '/html/body/div/div/main/div/div/div/div/form/div[3]/input')
        password_field.send_keys(password)
        submit_login_button = web.find_element("xpath", '/html/body/div/div/main/div/div/div/div/form/div[5]/input')
        submit_login_button.click()
        time.sleep(2)

        #navigate to reward codes page
        coupon_codes = web.find_element("xpath", '/html/body/div/div/main/div/div[1]/button[3]')
        coupon_codes.click()
        time.sleep(1)

        for i in range (row_start, sh.max_row+1):
           add_new = web.find_element("xpath", '/html/body/div/div/main/div/div[2]/div/div[1]/button')
           add_new.click()
           time.sleep(1)
           cell_obj = sh.cell(row=i, column=col_num)
           code_field = web.find_element("xpath", '/html/body/div/div/main/div/div[2]/div/div/form/div[1]/input')
           code_field.send_keys(cell_obj.value)
           submit_button = web.find_element("xpath", '/html/body/div/div/main/div/div[2]/div/div/form/div[8]/input')
           submit_button.click()
           time.sleep(1)
           st.write(f"Coupon code {cell_obj.value} has been added to the CMS system.")

        st.success("All coupon codes have been added to the CMS system")
